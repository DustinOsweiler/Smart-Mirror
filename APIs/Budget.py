import openpyxl 

path = str("C:/Users/dusti/Documents/Thesis/APIs/Budget.xlsx")

wb_obj = openpyxl.load_workbook(path, data_only=True) 

sheet_obj = wb_obj.active

maxrow = sheet_obj.max_row

TotalBudget = sheet_obj.cell(row = maxrow, column = 2)
TotalSpent = sheet_obj.cell(row = maxrow, column = 3)

TotalBudget = TotalBudget.value
TotalSpent = TotalSpent.value

#print(TotalBudget, TotalSpent)

if(TotalBudget < TotalSpent):
    output = "You are over budget!"
    left = "You don't have any left to spend in your budget"
else:
    output = "You are under budget!"
    left = "You have $" + str(TotalBudget - TotalSpent) + " left to spend this month"


def getTotal():
    path = str("C:/Users/dusti/Documents/Thesis/APIs/Budget.xlsx")

    wb_obj = openpyxl.load_workbook(path, data_only=True) 

    sheet_obj = wb_obj.active

    maxrow = sheet_obj.max_row

    TotalBudget = sheet_obj.cell(row = maxrow, column = 2)
    TotalSpent = sheet_obj.cell(row = maxrow, column = 3)

    TotalBudget = TotalBudget.value
    TotalSpent = TotalSpent.value

    #print(TotalBudget, TotalSpent)

    if(TotalBudget < TotalSpent):
        output = "You are over budget!"
        left = "You don't have any left to spend in your budget"
    else:
        output = "You are under budget!"
        left = "You have $" + str(TotalBudget - TotalSpent) + " left to spend this month"

    return(TotalBudget)

def getSpent():
    path = str("C:/Users/dusti/Documents/Thesis/APIs/Budget.xlsx")

    wb_obj = openpyxl.load_workbook(path, data_only=True) 

    sheet_obj = wb_obj.active

    maxrow = sheet_obj.max_row

    TotalBudget = sheet_obj.cell(row = maxrow, column = 2)
    TotalSpent = sheet_obj.cell(row = maxrow, column = 3)

    TotalBudget = TotalBudget.value
    TotalSpent = TotalSpent.value

    #print(TotalBudget, TotalSpent)

    if(TotalBudget < TotalSpent):
        output = "You are over budget!"
        left = "You don't have any left to spend in your budget"
    else:
        output = "You are under budget!"
        left = "You have $" + str(TotalBudget - TotalSpent) + " left to spend this month"

    return(TotalSpent)

def getOutput():
    path = str("C:/Users/dusti/Documents/Thesis/APIs/Budget.xlsx")

    wb_obj = openpyxl.load_workbook(path, data_only=True) 

    sheet_obj = wb_obj.active

    maxrow = sheet_obj.max_row

    TotalBudget = sheet_obj.cell(row = maxrow, column = 2)
    TotalSpent = sheet_obj.cell(row = maxrow, column = 3)

    TotalBudget = TotalBudget.value
    TotalSpent = TotalSpent.value

    #print(TotalBudget, TotalSpent)

    if(TotalBudget < TotalSpent):
        output = "You are over budget!"
        left = "You don't have any left to spend in your budget"
    else:
        output = "You are under budget!"
        left = "You have $" + str(TotalBudget - TotalSpent) + " left to spend this month"

    return(output)

def getLeft():
    path = str("C:/Users/dusti/Documents/Thesis/APIs/Budget.xlsx")

    wb_obj = openpyxl.load_workbook(path, data_only=True) 

    sheet_obj = wb_obj.active

    maxrow = sheet_obj.max_row

    TotalBudget = sheet_obj.cell(row = maxrow, column = 2)
    TotalSpent = sheet_obj.cell(row = maxrow, column = 3)

    TotalBudget = TotalBudget.value
    TotalSpent = TotalSpent.value

    #print(TotalBudget, TotalSpent)

    if(TotalBudget < TotalSpent):
        output = "You are over budget!"
        left = "You don't have any left to spend in your budget"
    else:
        output = "You are under budget!"
        left = "You have $" + str(TotalBudget - TotalSpent) + " left to spend this month"

    return(left)
    
    #print(output, left)